from django.shortcuts import render,redirect
from.models import user_register
from django.core.exceptions import ValidationError
from django.core.validators import EmailValidator
from django.contrib import messages
from .models import*
# Create your views here.
def indexpage(request):
    return render(request,'index.html')
def aboutpage(request):
    return render(request,'about.html')
def contactpage(request):
    return render(request,'contact.html')
def servicespage(request):
    return render(request,'services.html')

def logout(request):
    request.session.flush()
    return redirect('indexpage')
def registerpage(request):
    if request.method=="POST":
        name=request.POST.get('name')
        username=request.POST.get('username')
        phone=request.POST.get('phone')
        email=request.POST.get('email')
        image=request.FILES.get('image')
        password=request.POST.get('password')
        re_password=request.POST.get('re_password')
        if password !=re_password:
            return render(request,'register.html',{'error_msg':'Passwords do not match!'})
        if email != email.lower():
            return render(request,'register.html',{'error_msg':'email must be lowercase'})
        try:
            EmailValidator()(email)
        except ValidationError:
            return render(request,'register.html',{'error_msg':'Invalid email format'})
        if user_register.objects.filter(email=email):
            alert="<script>alert('Email already exist');window.location.href='/login/';</script>"
            return HttpResponse(alert)
        

        obj=user_register(Name=name,username=username,phone=phone,email=email,image=image,
                                 password=password,re_password=re_password)
        obj.save()
        return redirect('login')
    return render(request,'register.html')

from django.core.mail import send_mail
from django.conf import settings
import random

def loginpage(request):
    if request.method == "POST":
        # Step 1: Initial Login Phase
        if 'email' in request.POST and 'password' in request.POST:
            email = request.POST.get('email')
            password = request.POST.get('password')
            try:
                user = user_register.objects.get(email=email, password=password)
                # Generate OTP
                otp = random.randint(100000, 999999)
                request.session['otp'] = otp
                request.session['email'] = user.email

                # Send OTP via Email
                send_mail(
                    'Your Login OTP',
                    f'Your OTP is {otp}',
                    settings.EMAIL_HOST_USER,
                    [user.email],
                    fail_silently=False,
                )
                return render(request, 'verify_otp.html', {'msg': 'OTP sent to your email.'})
            except user_register.DoesNotExist:
                msg = "Invalid email or password"
                return render(request, 'login.html', {'msg': msg})

        # Step 2: OTP Verification Phase
        if 'otp' in request.POST:
            otp = request.POST.get('otp')
            if str(request.session.get('otp')) == otp:
                # OTP matches, log the user in
                return redirect('home')
            else:
                return render(request, 'verify_otp.html', {'msg': 'Invalid OTP'})

    return render(request, 'login.html')

def profile(request):
    if 'email' in  request.session:
        mail=request.session['email']
        usr=user_register.objects.get(email=mail)
    return render(request,'profile.html',{'usr':usr})

def edit(request,uid):
    edt=user_register.objects.get(id=uid)
    if request.method=="POST":
        name=request.POST.get('name')
        username=request.POST.get('username')
        phone=request.POST.get('phone')
        email=request.POST.get('email')
        image=request.FILES.get('image')
        edt.Name=name
        edt.username=username
        edt.phone=phone
        edt.email=email
        if image is not None:
            edt.image=image
        edt.save()
        return redirect('profile')
    return render(request,'edit.html',{'edt':edt})


def adminindex(request):
    return render(request,'adminhome.html')

def adminlogin(request):
    if request.method=="POST":
        uname=request.POST.get('username')
        passw=request.POST.get('password')
        u='admin'
        p='admin'
        if uname==u:
            if passw==p:
                return redirect('adminindex')
    return render(request,'adminlogin.html')

#userlist viewed by admin
def userlist(request):
    user=user_register.objects.all()
    return render(request,'userlist.html',{'user':user})

#user deleted by admin
def user_delete(request,did):
    user=user_register.objects.get(id=did)
    user.delete()
    return redirect('userlist')

from django.http import HttpResponse
from django.shortcuts import render
from .models import Feedback

from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import user_register, Feedback

def feedback(request):
    email = request.session.get('email')
    if not email:
        return redirect('/login')  # Redirect to login if the user is not logged in

    user = user_register.objects.filter(email=email).first()
    if request.method == "POST":
        feedback_text = request.POST.get('feedback_text')
        rating = request.POST.get('rating')

        # Check for missing fields
        if not feedback_text or not rating:
            return HttpResponse("<script>alert('Please fill in all required fields.'); window.location.href='/feedback_rate';</script>")

        try:
            rating = int(rating)
            if rating not in [1, 2, 3, 4, 5]:
                raise ValueError("Invalid rating value")
        except (ValueError, TypeError):
            return HttpResponse("<script>alert('Invalid rating value. Please select a valid rating.'); window.location.href='/feedback_rate';</script>")

        # Save feedback to the database
        Feedback.objects.create(
            feedback_text=feedback_text,
            rating=rating,
            email=email  # Save the email from the session
        )

        return HttpResponse("<script>alert('Feedback submitted successfully!'); window.location.href='/home';</script>")

    # Render the feedback form for GET requests
    return render(request, 'feedback_rate.html', {'e': user})

    
def feedbacklist(request):
    f=Feedback.objects.all()
    return render(request,'feedbacklist.html',{'f':f})

def deletefeedback(request,id):
    f=Feedback.objects.get(id=id)
    f.delete()
    return redirect('feedbacklist')


def home(request):
    email = request.session.get('email')
    user = user_register.objects.filter(email=email).first()
    return render(request,'home.html',{'user':user})

from datetime import datetime
from django.shortcuts import render, redirect
from .models import Income, Expense, user_register
from django.db.models import Sum
from decimal import Decimal

from datetime import datetime
from decimal import Decimal

from datetime import datetime

from datetime import datetime

from datetime import datetime
from django.shortcuts import render, redirect
from .models import Income, Expense, user_register

def addincome(request):
    email = request.session.get('email')
    user = user_register.objects.filter(email=email).first()

    if email and user:
        if request.method == "POST":
            incomesource = request.POST.get('incomesource')
            income = int(request.POST.get('income'))  # Ensure income is treated as an integer
            date = request.POST.get('date')

            # Get the month and year from the date
            income_date = datetime.strptime(date, '%Y-%m-%d')
            income_month = income_date.month
            income_year = income_date.year

            # Get all income records for the current month
            all_incomes_this_month = Income.objects.filter(user=user, date__month=income_month, date__year=income_year)

            # Calculate the total income by summing all previous incomes for this month
            total_income_for_user = sum([income_record.income for income_record in all_incomes_this_month]) + income

            # Get all expenses for the current month
            all_expenses_this_month = Expense.objects.filter(user=user, date__month=income_month, date__year=income_year)

            # Calculate the total expenses for the month
            total_expenses_for_user = sum([expense.amount for expense in all_expenses_this_month])

            # Find the most recent income record (the last added income for the current month)
            last_income_record = all_incomes_this_month.order_by('-id').first()

            if last_income_record:
                # If there is a previous record, increment the balance by the new income, subtracting the total expenses
                balance = last_income_record.balance + income
            else:
                # If no previous records, set balance as the new income minus the total expenses
                balance = income - total_expenses_for_user

            # Create the new income record with the updated totalincome and balance
            Income.objects.create(
                incomesource=incomesource,
                income=income,
                totalincome=total_income_for_user,  # Updated total income
                date=date,
                user=user,
                balance=balance  # Updated balance after adding the new income
            )
            return HttpResponse("<script>alert('income added successfully!'); window.location.href='/home';</script>")
            return redirect('home')

        return render(request, 'addincome.html', {'user': user})

    return render(request, 'addincome.html', {'user': user})






def list_income(request):
    email = request.session.get('email')
    if email:
        u=user_register.objects.get(email=email)
        income = Income.objects.filter(user=u)
        return render(request, 'listincome.html', {'income': income,'u': u})
    

from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Income, user_register
from datetime import datetime

from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Income, user_register
from datetime import datetime

def edit_income(request, income_id):
    # Fetch the user based on email stored in the session
    email = request.session.get('email')
    user = user_register.objects.filter(email=email).first()

    if email and user:
        # Fetch the income record that needs to be edited
        income = Income.objects.filter(id=income_id, user=user).first()

        if not income:
            messages.error(request, "Income record not found.")
            return redirect('home')

        if request.method == "POST":
            # Fetch the updated data
            new_income = int(request.POST.get('income'))  # Ensure income is treated as an integer
            date = request.POST.get('date')
            income_source = request.POST.get('incomesource')

            # Calculate the difference between the new income and the old income
            income_difference = new_income - income.income

            # Update the income record with new values
            income.income = new_income
            income.date = date
            income.incomesource = income_source

            # Save the updated income record
            income.save()

            # Get all income records for the same month and year
            income_date = datetime.strptime(date, '%Y-%m-%d')
            income_month = income_date.month
            income_year = income_date.year

            all_incomes_this_month = Income.objects.filter(user=user, date__month=income_month, date__year=income_year)

            # Recalculate total income for the current month
            total_income_for_user = sum([i.income for i in all_incomes_this_month])

            # Update balance for all incomes in the current month
            balance = 0  # Start from zero
            for income_record in all_incomes_this_month:
                balance += income_record.income  # Increase balance with the new incomes
                income_record.totalincome = total_income_for_user  # Update total income
                income_record.balance = balance  # Update balance with the latest total income
                income_record.save()

            # Redirect back to the homepage with success message
            messages.success(request, "Income updated successfully.")
            return redirect('listincome')

        # Pre-populate the form with existing income data for editing
        return render(request, 'edit_income.html', {'user': user, 'income': income})

    messages.error(request, "You need to be logged in.")
    return redirect('login')

from django.db.models import Sum

def deleteincome(request, in_id):
    # Fetch the income record to be deleted
    income = Income.objects.filter(id=in_id).first()

    if not income:
        messages.error(request, "Income record not found.")
        return redirect('home')

    # Capture key data before deleting the record
    income_month = income.date.month
    income_year = income.date.year
    user = income.user

    # Fetch total expenses for the same user and month
    total_expenses = Expense.objects.filter(
        user=user,
        date__month=income_month,
        date__year=income_year
    ).aggregate(total_expense=Sum('amount'))['total_expense'] or 0  # Default to 0 if no expenses

    # Delete the income record
    income.delete()

    # Fetch all remaining incomes for the same user and month
    all_incomes_this_month = Income.objects.filter(
        user=user,
        date__month=income_month,
        date__year=income_year
    ).order_by('date', 'id')

    # Recalculate total income and balance after deletion
    total_income_so_far = 0
    for income_record in all_incomes_this_month:
        total_income_so_far += income_record.income  # Accumulate incomes
        income_record.totalincome = total_income_so_far  # Update total income
        # Correctly calculate the balance as total income minus total expenses
        income_record.balance = total_income_so_far - total_expenses
        income_record.save()

    messages.success(request, "Income deleted and totals updated successfully.")
    return redirect('listincome')








from decimal import Decimal
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Income, Expense, user_register
from datetime import datetime

from decimal import Decimal
from datetime import datetime
from django.contrib import messages

def addexpense(request):
    email = request.session.get('email')
    user = user_register.objects.filter(email=email).first()

    if user:
        if request.method == "POST":
            # Fetch form data
            expensecategory = request.POST.get('expensecategory')
            amount = request.POST.get('amount')
            date = request.POST.get('date')
            income_id = request.POST.get('income_id')

            # Input Validation
            if not expensecategory or not amount or not date or not income_id:
                messages.error(request, "All fields are required.")
                return redirect('addexpense')

            try:
                amount = Decimal(amount)  # Convert amount to Decimal
            except Exception:
                messages.error(request, "Invalid amount format.")
                return redirect('addexpense')

            # Validate and fetch income record for the correct month and year
            income = Income.objects.filter(id=income_id, user=user).first()
            if not income:
                messages.error(request, "Income record not found.")
                return redirect('addexpense')

            # Ensure balance is initialized if not already set
            if income.balance is None or income.balance == 0:
                income.balance = Decimal(income.income)  # Set balance equal to the income value initially
                income.save()

            # Save the expense record
            expense = Expense(
                expensecategory=expensecategory,
                amount=amount,
                date=date,
                user=user
            )
            expense.save()

            # Check for the month of the expense
            expense_date = datetime.strptime(date, '%Y-%m-%d')  # Convert string date to datetime object
            expense_month = expense_date.month
            expense_year = expense_date.year

            # Find the latest income for the same month and year by ID (last added income)
            latest_income = Income.objects.filter(user=user, date__month=expense_month, date__year=expense_year).order_by('-id').first()

            if latest_income:
                # Subtract the expense amount from the last added income's balance
                latest_income.balance -= amount  # Deduct the expense from the balance
                latest_income.save()

                # Notify user if their balance is now negative
                if latest_income.balance < 0:
                    messages.warning(request, 'Financial Alert: Expenses Exceed Income for this month.')
            messages.warning(request, 'Expense added successfully!')
            return redirect('expense_limit')
            return redirect('addexpense')

        # Preload income data for rendering the form
        income = Income.objects.filter(user=user).first()
        return render(request, 'addexpense.html', {'user': user, 'income': income})

    messages.error(request, "You need to be logged in.")
    return redirect('login')



from django.shortcuts import render, redirect
from django.contrib import messages
from decimal import Decimal
from datetime import datetime
from .models import Expense, Income, user_register

# def editexpense(request, expense_id):
#     email = request.session.get('email')
#     user = user_register.objects.filter(email=email).first()

#     if user:
#         expense = Expense.objects.filter(id=expense_id, user=user).first()
#         if not expense:
#             messages.error(request, "Expense record not found.")
#             return redirect('expenseadd')

#         if request.method == "POST":
#             # Fetch form data
#             expensecategory = request.POST.get('expensecategory')
#             amount = request.POST.get('amount')
#             date = request.POST.get('date')
            
#             # Input Validation
#             if not expensecategory or not amount or not date:
#                 messages.error(request, "All fields are required.")
#                 return redirect('editexpense', expense_id=expense.id)

#             try:
#                 amount = Decimal(amount)  # Convert amount to Decimal
#             except Exception:
#                 messages.error(request, "Invalid amount format.")
#                 return redirect('editexpense', expense_id=expense.id)

#             # Retrieve the related income record for the same month and year
#             income = Income.objects.filter(id=expense.income.id).first()
#             if not income:
#                 messages.error(request, "Related income record not found.")
#                 return redirect('editexpense', expense_id=expense.id)

#             # Deduct the old expense from the income balance
#             income.balance += expense.amount  # Revert the old expense amount
#             income.save()

#             # Update the expense record with the new data
#             expense.expensecategory = expensecategory
#             expense.amount = amount
#             expense.date = date
#             expense.save()

#             # Subtract the new expense from the income balance
#             income.balance -= amount
#             income.save()

#             # Notify user if their balance is now negative
#             if income.balance < 0:
#                 messages.warning(request, 'Financial Alert: Expenses Exceed Income for this month.')


#             messages.success(request, "Expense updated successfully.")
#             return redirect('expenselist')

#         return render(request, 'editexpense.html', {'user': user, 'expense': expense})

#     messages.error(request, "You need to be logged in.")
#     return redirect('login')



def expenselist(request):
    email=request.session.get('email')
    if email:
        u=user_register.objects.get(email=email)
        ex=Expense.objects.filter(user=u)

        return render (request, 'expenselist.html',{'e':ex})

from django.shortcuts import render, redirect
from .models import user_register, Expense
from datetime import datetime

from django.shortcuts import get_object_or_404, redirect
from .models import Expense, Income

from decimal import Decimal

from decimal import Decimal
from django.db import transaction


def editexpense(request, expense_id):
    email = request.session.get('email')
    if email:
        u = user_register.objects.get(email=email)
        expense = get_object_or_404(Expense, id=expense_id)

        # Fetch the latest income record for the same month and year as the expense
        expense_date = expense.date  # Already a datetime.date object
        income = Income.objects.filter(
            user=u,
            date__year=expense_date.year,
            date__month=expense_date.month
        ).order_by('-id').first()

        if not income:
            messages.error(request, "No income record found for the month and year of the expense.")
            return redirect('expenselist')  # Redirect to expense list or appropriate page

        if request.method == 'POST':
            # Get the new expense values from the POST data
            try:
                new_expense_amount = Decimal(request.POST.get('amount'))  # Convert to Decimal
                old_expense_amount = Decimal(expense.amount)  # Ensure it's a Decimal
            except Exception as e:
                messages.error(request, "Invalid amount format.")
                return redirect('editexpense', expense_id=expense_id)

            # Update Expense fields
            
            expense.amount = new_expense_amount
            expense.date = request.POST.get('date', expense.date)
            expense.expensecategory = request.POST.get('expensecategory', expense.expensecategory)

            # Use transaction.atomic() to ensure both expense and income are updated together
            try:
                with transaction.atomic():
                    # Save the updated expense
                    expense.save()

                    # Adjust the income balance
                    income_balance = Decimal(income.balance)
                    income.balance = income_balance + old_expense_amount - new_expense_amount
                    income.save()

                    # Debugging: Check the updated balance
                    print(f"Updated Income Balance: {income.balance}")

                    # Redirect after saving the update
                    messages.success(request, "Expense updated successfully.")
                    return redirect('expenselist')  # Redirect to expense list or desired page
            except Exception as e:
                # If there's an error, log the exception
                print(f"Error updating expense and income: {e}")
                messages.error(request, "An error occurred while updating the expense.")

        # Render the edit expense page if GET request
        return render(request, 'editexpense.html', {'u': u, 'edit': expense, 'income': income})

    messages.error(request, "You need to be logged in.")
    return redirect('login')







import matplotlib
matplotlib.use('Agg')  # Use the Agg backend (no GUI interaction)

import matplotlib.pyplot as plt
from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import render
from .models import Expense, user_register
import base64
import calendar  # For converting month numbers to month names
import numpy as np  # For color mapping

def expense_bar_chart(request):
    email = request.session.get('email')  # This assumes the email is stored in the session after login

    if not email:
        return HttpResponse("User not logged in", status=401)

    try:
        user = user_register.objects.get(email=email)
    except user_register.DoesNotExist:
        return HttpResponse("User not found", status=404)

    expenses = Expense.objects.filter(user=user)

    # Get distinct year-month combinations
    months = expenses.values('date__year', 'date__month').distinct()

    bar_chart_images = []
    for month in months:
        # Filter expenses for the current month
        month_expenses = expenses.filter(date__year=month['date__year'], date__month=month['date__month'])

        # Aggregate expenses by day and category
        daily_totals = {}
        category_labels = {}
        category_colors = {}

        unique_categories = set(expenses.values_list('expensecategory', flat=True))  # Get all unique categories
        cmap = plt.get_cmap("tab10", len(unique_categories))  # Get a colormap with unique colors
        category_color_map = {category: cmap(i) for i, category in enumerate(unique_categories)}

        for expense in month_expenses:
            day = expense.date.day
            category = expense.expensecategory

            daily_totals.setdefault(day, {}).setdefault(category, 0)
            daily_totals[day][category] += expense.amount
            category_labels.setdefault(day, set()).add(category)

        # Prepare labels and values for the bar chart
        days = sorted(daily_totals.keys())  # Sort days in order
        category_stack = {cat: [] for cat in unique_categories}  # Empty lists for stacking
        bottom_stack = np.zeros(len(days))  # Start bottom at zero

        # Create the stacked bar data
        for i, day in enumerate(days):
            for category in unique_categories:
                amount = daily_totals[day].get(category, 0)
                category_stack[category].append(amount)

        # Convert the month number to its name
        month_name = calendar.month_name[month['date__month']]

        # Dynamically adjust figure size based on number of days
        fig_width = max(8, len(days) * 0.5)
        plt.figure(figsize=(fig_width, 5))

        # Create the stacked bar chart
        for category, values in category_stack.items():
            bars = plt.bar(days, values, color=category_color_map[category], label=category, bottom=bottom_stack)
            bottom_stack += np.array(values)

        plt.xlabel('Day of the Month')
        plt.ylabel('Amount Spent')
        plt.title(f"Daily Expense Distribution for {user.Name} ({month_name} {month['date__year']})")

        # Add legend for categories
        plt.legend(title="Expense Categories", loc='upper right', fontsize=8)

        # Adjust layout
        plt.xticks(days)  # Ensure all days are shown
        plt.tight_layout()

        # Save the bar chart to a BytesIO object
        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        plt.close()

        # Convert the image to base64 to embed in HTML
        img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        bar_chart_images.append(f"data:image/png;base64,{img_base64}")

    # Render the template with the list of bar chart images
    return render(request, 'expense_chart.html', {'bar_chart_images': bar_chart_images})






from .models import Expenseimage
def addexpenseimage(request):
    email=request.session.get('email')
    user = user_register.objects.filter(email=email).first()
    if user:
        
        if request.method =="POST":
            image = request.FILES.get('image')
            Expenseimage(image=image,user=user).save()
            return redirect('home')
        return render(request,'expenseimage.html',{'u':user})
    return render(request,'expenseimage.html',{'u':user})
    


#voice 

from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
import re
from datetime import datetime
from .models import Expense, user_register

@csrf_exempt
def add_expense_voice(request):
    if request.method == 'GET':
        return render(request, 'addexpensevoice.html')

    elif request.method == 'POST':
        try:
            # Parse the incoming request data
            data = json.loads(request.body)
            recognized_text = data.get('recognized_text', '').strip()
            print("Received text:", recognized_text)  # Debugging output

            email = request.session.get('email')

            if not email:
                return JsonResponse({'status': 'error', 'message': 'User email is missing.'})

            user = user_register.objects.get(email=email)

            # Improved Regex to match "amount category YYYY Month DD"
            # pattern = r'^\s*(\d+)\s+([a-zA-Z]+)\s+(\d{4})\s+([a-zA-Z]+)\s+(\d{1,2})\s*$'
            pattern = r'^\s*(\d+)\s+([a-zA-Z\s]+)\s+(\d{4})\s+([a-zA-Z]+)\s+(\d{1,2})\s*$'



            match = re.match(pattern, recognized_text, re.IGNORECASE)

            if match:
                # Extracting the details from the matched pattern
                amount = int(match.group(1))
                category = match.group(2).capitalize()  # Capitalize the category
                year = int(match.group(3))
                month = match.group(4).capitalize()  # Capitalize the month
                day = int(match.group(5))

                # Convert month name to number
                try:
                    month_number = datetime.strptime(month, "%B").month
                except ValueError as e:
                    print(f"Month Conversion Error: {e}")
                    return JsonResponse({'status': 'error', 'message': f'Invalid month: {month}. Please check the month spelling.'})

                # Format date as YYYY-MM-DD
                formatted_date = f"{year}-{month_number:02d}-{day:02d}"

                # Find the corresponding income record
                income = Income.objects.filter(user=user, date__year=year, date__month=month_number).order_by('-id').first()

                if not income:
                    return JsonResponse({'status': 'error', 'message': 'No income record found for the specified date.'})

                # Ensure balance is initialized if not already set
                if income.balance is None or income.balance == 0:
                    income.balance = Decimal(income.income)  # Set balance equal to the income value initially
                    income.save()

                # Save the expense to the database with the user association
                expense = Expense(
                    user=user,  # Associate with the user
                    expensecategory=category,
                    amount=amount,
                    date=formatted_date
                )
                expense.save()

                # Subtract the expense amount from the income balance
                income.balance -= Decimal(amount)
                income.save()

                # Notify the user if the balance is now negative
                if income.balance < 0:
                    print("Financial Alert: Expenses exceed income for the month.")

                print("Expense saved successfully.")
                return JsonResponse({
                    'status': 'success',
                    'message': f'Expense added: {amount} to {category} for {formatted_date}. Remaining balance: {income.balance}'
                })
            else:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Received data format is incorrect. Please try again.'
                })
        except Exception as e:
            # Log the full exception for debugging
            print(f"Exception occurred: {e}")
            return JsonResponse({
                'status': 'error',
                'message': 'An error occurred while processing your expense.'
            })

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})


# from django.shortcuts import render
# from django.db.models import Avg
# from .models import Expense

# def predict_expenses(request):
#     categories = Expense.objects.values_list('expensecategory', flat=True).distinct()
#     selected_category = request.GET.get('category')
#     selected_date = request.GET.get('date')
#     prediction_result = None
#     warning_message = None

#     if selected_category and selected_date:
#         # Calculate prediction based on the average expense for the category
#         average_expense = Expense.objects.filter(
#             expensecategory=selected_category
#         ).aggregate(avg_amount=Avg('amount'))['avg_amount']

#         if average_expense:
#             prediction_result = {
#                 "category": selected_category,
#                 "date": selected_date,
#                 "predicted_amount": round(average_expense, 2),
#             }
#         else:
#             warning_message = f"No data available for the category '{selected_category}'."

#     return render(request, 'predict.html', {
#         "categories": categories,
#         "selected_category": selected_category,
#         "selected_date": selected_date,
#         "prediction_result": prediction_result,
#         "warning_message": warning_message,
#     })
from django.shortcuts import render
from django.db.models import Avg
from .models import Expense
from datetime import datetime
import calendar

def predict_expenses(request):
    categories = Expense.objects.values_list('expensecategory', flat=True).distinct()
    selected_category = request.GET.get('category')
    selected_month = request.GET.get('month')  # Expecting format like '2025-03'
    prediction_result = None
    warning_message = None

    if selected_category and selected_month:
        try:
            # Convert selected_month to datetime object
            prediction_date = datetime.strptime(selected_month, '%Y-%m')
            month_name = calendar.month_name[prediction_date.month]
            year = prediction_date.year

            # Calculate average monthly expense for the selected category
            # Group by month and year, then average those monthly totals
            monthly_averages = (Expense.objects
                              .filter(expensecategory=selected_category)
                              .extra(select={'month': "strftime('%%Y-%%m', date)"})
                              .values('month')
                              .annotate(monthly_total=Avg('amount'))
                              .aggregate(avg_monthly=Avg('monthly_total'))['avg_monthly'])

            if monthly_averages:
                prediction_result = {
                    "category": selected_category,
                    "month": f"{month_name} {year}",
                    "predicted_amount": round(monthly_averages, 2),
                }
            else:
                warning_message = f"No data available for the category '{selected_category}'."

        except ValueError:
            warning_message = "Invalid month format. Please use YYYY-MM format."

    return render(request, 'predict.html', {
        "categories": categories,
        "selected_category": selected_category,
        "selected_month": selected_month,
        "prediction_result": prediction_result,
        "warning_message": warning_message,
    })

from .models import ExpenseLimit, Expense
from django.shortcuts import render
from django.db.models import Sum
from datetime import datetime

def set_category_limit(request):
    email = request.session.get('email')
    user = user_register.objects.get(email=email)
    message = ""

    if request.method == 'POST':
        category = request.POST.get('category')
        limit_amount = request.POST.get('limit')

        try:
            # Convert the limit amount to a float
            limit_amount = float(limit_amount)
            # Create or update the limit for the given category
            expense_limit, created = ExpenseLimit.objects.update_or_create(
                user=user, expensecategory=category,
                defaults={'limit_amount': limit_amount}
            )
            message = f"Limit of {limit_amount} has been set successfully for category '{category}'."
        except ValueError:
            message = "Invalid limit value. Please enter a valid number."

    # Fetch all distinct categories for display
    categories = Expense.objects.values_list('expensecategory', flat=True).distinct()
    expense=ExpenseLimit.objects.filter(user=user)
    return render(request, 'addlimit.html', {
        'categories': categories,
        'message': message,
        'expense':expense
    })
def delete_expense_limit(request, limit_id):
    expense_limit = get_object_or_404(ExpenseLimit, id=limit_id)
    expense_limit.delete()
    messages.success(request, "Expense limit deleted successfully!")
    return redirect('set_expense_limit') 
# Redirect to expense limit list page

from datetime import datetime
from django.shortcuts import render
from .models import user_register, ExpenseLimit, Expense  # Ensure these models exist

def category_limit_check(request):
    email = request.session.get('email')
    user = user_register.objects.get(email=email)

    # Get the current month and year
    current_month = datetime.now().month
    current_year = datetime.now().year

    # Fetch all expense limits set by the user
    limits = ExpenseLimit.objects.filter(user=user)

    category_status = []

    # Loop through each category limit
    for limit in limits:
        category = limit.expensecategory
        limit_amount = float(limit.limit_amount) if limit.limit_amount is not None else 0.0  # Ensure float type

        # Initialize total expense to 0
        total_expense = 0.0

        # Fetch expenses for the current month and category
        expenses = Expense.objects.filter(
            user=user,
            expensecategory=category,
            date__month=current_month,
            date__year=current_year
        )

        # Debugging information
        print(f"Checking category: {category}")
        print(f"Querying expenses for Month: {current_month}, Year: {current_year}")
        print("Expense Query:", expenses.query)

        # Calculate total expenses
        if expenses.exists():
            total_expense = sum(exp.amount for exp in expenses if exp.amount is not None)

        # Ensure `total_expense` is float
        total_expense = float(total_expense)

        print(f"Total Expenses for {category}: {total_expense}, Limit: {limit_amount}")

        # Debugging before comparison
        print(f"DEBUG: Comparing Total Expense ({total_expense}) with Limit ({limit_amount})")
        print(f"DEBUG: Limit Exceeded? {total_expense > limit_amount}")

        # Check if the expense limit is exceeded
        limit_exceeded = total_expense > limit_amount

        category_status.append({
            'category': category,
            'limit_amount': limit_amount,
            'total_expense': total_expense,
            'limit_exceeded': limit_exceeded,
            'error_message': (
                f"Expense limit for {category} exceeded! Total Expense: ₹{total_expense}, Limit: ₹{limit_amount}."
                if limit_exceeded else None
            ),
            'user': user
        })

    return render(request, 'expense_limit.html', {'category_status': category_status})

import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.shortcuts import render
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from .models import Income, Expense, user_register
from datetime import datetime

def export_report(request):
    user_email = request.session.get('email')
    if not user_email:
        return HttpResponse("No user email found in session.", status=400)
    
    user = user_register.objects.get(email=user_email)

    if request.method == 'POST':
        # Get form data
        try:
            month = int(request.POST.get('month'))
            year = int(request.POST.get('year'))
            download_type = request.POST.get('type')
        except ValueError:
            return render(request, 'export_report.html', {
                'current_year': datetime.now().year,
                'error': 'Invalid month or year input.'
            })

        # Validate inputs
        if month < 1 or month > 12 or year < 2000 or year > 2100:
            return render(request, 'export_report.html', {
                'current_year': datetime.now().year,
                'error': 'Month must be between 1-12 and year between 2000-2100.'
            })

        # Filter data for the logged-in user and specific month/year
        user_incomes = Income.objects.filter(
            user=user,
            date__year=year,
            date__month=month
        )
        user_expenses = Expense.objects.filter(
            user=user,
            date__year=year,
            date__month=month
        )

        # Month name for display purposes
        month_name = datetime(year, month, 1).strftime('%B')

        # Excel Generation
        if download_type == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{month_name} {year} Report"
            ws.append([f"Income and Expense Report for {month_name} {year}"])
            ws.append([])  # Blank row
            ws.append(["User Email", "Income Source", "Income", "Total Income", "Date", "Balance"])  # Income Header

            # Write Income data
            for income in user_incomes:
                ws.append([
                    income.user.email,
                    income.incomesource,
                    income.income,
                    income.totalincome,
                    income.date,
                    income.balance
                ])

            # Add a blank row before Expense data
            ws.append([])
            ws.append(["User Email", "Expense Category", "Amount", "Date"])  # Expense Header

            # Write Expense data
            for expense in user_expenses:
                ws.append([
                    expense.user.email,
                    expense.expensecategory,
                    expense.amount,
                    expense.date
                ])
            
            # Style the headers
            for cell in ws[3] + ws[len(user_incomes) + 5]:
                cell.font = Font(bold=True)
            
            # Save Excel file to memory
            excel_response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            excel_response['Content-Disposition'] = f'attachment; filename="Income_Expense_{month_name}_{year}.xlsx"'
            wb.save(excel_response)
            return excel_response

        # PDF Generation
        elif download_type == 'pdf':
            pdf_response = HttpResponse(content_type='application/pdf')
            pdf_response['Content-Disposition'] = f'attachment; filename="Income_Expense_{month_name}_{year}.pdf"'
            
            doc = SimpleDocTemplate(pdf_response, pagesize=letter)
            elements = []
            
            # Add a title
            elements.append(Table(
                [[f"Income and Expense Report for {user.Name} - {month_name} {year}"]],
                style=[('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                       ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                       ('FONTSIZE', (0, 0), (-1, -1), 14),
                       ('BOTTOMPADDING', (0, 0), (-1, -1), 12)]
            ))

            # Income Table
            if user_incomes.exists():
                income_data = [["Source", "Amount", "Total Income", "Date", "Balance"]]
                for income in user_incomes:
                    income_data.append([
                        income.incomesource,
                        income.income,
                        income.totalincome,
                        income.date,
                        income.balance
                    ])
                income_table = Table(income_data, colWidths=[100, 60, 80, 80, 80])
                income_table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                ]))
                elements.append(Table([[f"Income Details for {month_name} {year}"]], style=[('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                                                                                           ('FONTSIZE', (0, 0), (-1, -1), 12),
                                                                                           ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold')]))
                elements.append(income_table)

            # Expense Table
            if user_expenses.exists():
                expense_data = [["Category", "Amount", "Date"]]
                for expense in user_expenses:
                    expense_data.append([
                        expense.expensecategory,
                        expense.amount,
                        expense.date
                    ])
                expense_table = Table(expense_data, colWidths=[150, 80, 80])
                expense_table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                ]))
                elements.append(Table([[f"Expense Details for {month_name} {year}"]], style=[('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                                                                                             ('FONTSIZE', (0, 0), (-1, -1), 12),
                                                                                             ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold')]))
                elements.append(expense_table)
            
            doc.build(elements)
            return pdf_response

    # GET request: Render the form
    return render(request, 'export_report.html', {
        'current_year': datetime.now().year
    })


# import openpyxl
# from openpyxl.styles import Font
# from django.http import HttpResponse
# from reportlab.lib.pagesizes import letter
# from reportlab.pdfgen import canvas
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
# from reportlab.lib import colors
# from .models import Income, Expense, user_register
# from .models import*

# def export_to_excel(request):
#     user_email = request.session.get('email')
#     if not user_email:
#         return HttpResponse("No user email found in session.", status=400)
#     user = user_register.objects.get(email=user_email)
    
#     # Filter data for the logged-in user
#     user_incomes = Income.objects.filter(user=user)
#     user_expenses = Expense.objects.filter(user=user)

#     # Create Excel workbook
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Income and Expense"
#     ws.append(["User Email", "Income Source", "Income", "Total Income", "Date", "Balance"])  # Income Header row

#     # Write Income data
#     for income in user_incomes:
#         ws.append([
#             income.user.email,
#             income.incomesource,
#             income.income,
#             income.totalincome,
#             income.date,
#             income.balance
#         ])

#     # Add a blank row before Expense data
#     ws.append([])
#     ws.append(["User Email", "Expense Category", "Amount", "Date"])  # Expense Header row

#     # Write Expense data
#     for expense in user_expenses:
#         ws.append([
#             expense.user.email,
#             expense.expensecategory,
#             expense.amount,
#             expense.date
#         ])
    
#     # Style the headers
#     for cell in ws[1] + ws[len(user_incomes) + 3]:
#         cell.font = Font(bold=True)
    
#     # Save Excel file to memory
#     excel_response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     excel_response['Content-Disposition'] = 'attachment; filename="Income_Expense_Report.xlsx"'
#     wb.save(excel_response)

#     # Create PDF with a table format
#     pdf_response = HttpResponse(content_type='application/pdf')
#     pdf_response['Content-Disposition'] = 'attachment; filename="Income_Expense_Report.pdf"'
    
#     doc = SimpleDocTemplate(pdf_response, pagesize=letter)
#     elements = []
    
#     # Add a title
#     elements.append(Table(
#         [[f"Income and Expense Report for {user.Name}"]],
#         style=[('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
#                ('FONTSIZE', (0, 0), (-1, -1), 14),
#                ('BOTTOMPADDING', (0, 0), (-1, -1), 12)]
#     ))

#     # Income Table
#     if user_incomes.exists():
#         income_data = [["Source", "Amount", "Total Income", "Date", "Balance"]]
#         for income in user_incomes:
#             income_data.append([
#                 income.incomesource,
#                 income.income,
#                 income.totalincome,
#                 income.date,
#                 income.balance
#             ])
#         income_table = Table(income_data, colWidths=[100, 60, 80, 80, 80])
#         income_table.setStyle(TableStyle([
#             ('GRID', (0, 0), (-1, -1), 1, colors.black),
#             ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
#             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#             ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
#             ('FONTSIZE', (0, 0), (-1, -1), 10),
#         ]))
#         elements.append(Table([[f"Income Details"]], style=[('ALIGN', (0, 0), (-1, -1), 'LEFT'),
#                                                              ('FONTSIZE', (0, 0), (-1, -1), 12),
#                                                              ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold')]))
#         elements.append(income_table)

#     # Expense Table
#     if user_expenses.exists():
#         expense_data = [["Category", "Amount", "Date"]]
#         for expense in user_expenses:
#             expense_data.append([
#                 expense.expensecategory,
#                 expense.amount,
#                 expense.date
#             ])
#         expense_table = Table(expense_data, colWidths=[150, 80, 80])
#         expense_table.setStyle(TableStyle([
#             ('GRID', (0, 0), (-1, -1), 1, colors.black),
#             ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
#             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#             ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
#             ('FONTSIZE', (0, 0), (-1, -1), 10),
#         ]))
#         elements.append(Table([[f"Expense Details"]], style=[('ALIGN', (0, 0), (-1, -1), 'LEFT'),
#                                                               ('FONTSIZE', (0, 0), (-1, -1), 12),
#                                                               ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold')]))
#         elements.append(expense_table)
    
#     doc.build(elements)

#     # Return Excel or PDF based on user choice
#     download_type = request.GET.get('type', 'excel')
#     if download_type == 'pdf':
#         return pdf_response
#     return excel_response

# from decimal import Decimal
# from datetime import date
# from django.db.models import Sum

# def predict_savings(request):
#     email = request.session.get('email')
#     if not email:
#         messages.error(request, "You need to be logged in.")
#         return redirect('login')

#     user = user_register.objects.get(email=email)

#     if request.method == 'POST':
#         # Get the goal data from POST
#         goal_name = request.POST.get('goal_name')
#         target_amount = request.POST.get('target_amount')
#         start_date = request.POST.get('start_date')

#         if not goal_name or not target_amount:
#             messages.error(request, "Please provide a goal name and target amount.")
#             return redirect('predict_savings')

#         try:
#             target_amount = Decimal(target_amount)  # Convert to Decimal
#         except Exception as e:
#             messages.error(request, "Invalid target amount.")
#             return redirect('predict_savings')

#         # Create the goal
#         goal = Goal.objects.create(
#             user=user,
#             name=goal_name,
#             target_amount=target_amount,
#             start_date=start_date
#         )

#         # Get selected categories for prediction (comma-separated in the input)
#         selected_categories = request.POST.get('selected_categories', '').split(',')
#         selected_categories = [category.strip() for category in selected_categories]

#         if not selected_categories:
#             messages.error(request, "Please select at least one category for prediction.")
#             return redirect('predict_savings')

#         # Calculate predicted savings based on all categories excluding selected ones
#         today = date.today()
#         start_date = today.replace(month=today.month - 3 if today.month > 3 else 12, day=1)  # Last 3 months

#         # Filter out selected categories and get expenses from the remaining categories
#         expenses = Expense.objects.filter(
#             user=user,
#         ).exclude(expensecategory__in=selected_categories, date__gte=start_date)

#         # Calculate total expenses for the remaining categories over the last 3 months
#         total_expenses = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
#         avg_monthly_expense = total_expenses / 3 if total_expenses > 0 else 0

#         # Convert total_expenses and avg_monthly_expense to Decimal
#         total_expenses = Decimal(total_expenses)
#         avg_monthly_expense = Decimal(avg_monthly_expense)

#         # Predict savings as a percentage of average expenses (e.g., 10%)
#         predicted_savings = avg_monthly_expense * Decimal(0.1)  # Saving 10% of average monthly expenses

#         # Convert goal's current savings to Decimal (if any)
#         current_savings = Decimal(goal.current_savings) if goal.current_savings else Decimal(0)

#         # Calculate remaining amount to save
#         remaining_to_save = target_amount - current_savings

#         # Render the goal and prediction information
#         return render(request, 'predict_savings.html', {
#             'goal': goal,
#             'predicted_savings': round(predicted_savings, 2),
#             'remaining_to_save': round(remaining_to_save, 2),
#             'total_expenses': round(total_expenses, 2),
#             'avg_monthly_expense': round(avg_monthly_expense, 2),
#         })

#     # Render the page with the form to set a goal
#     return render(request, 'predict_savings.html')

from decimal import Decimal
from datetime import date
from django.db.models import Sum
from django.shortcuts import render, redirect
from django.contrib import messages
from dateutil.relativedelta import relativedelta  # For calculating months between dates

def predict_savings(request):
    email = request.session.get('email')
    if not email:
        messages.error(request, "You need to be logged in.")
        return redirect('login')

    user = user_register.objects.get(email=email)

    if request.method == 'POST':
        # Get the goal data from POST
        goal_name = request.POST.get('goal_name')
        target_amount = request.POST.get('target_amount')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        selected_categories = request.POST.get('selected_categories', '').split(',')

        if not goal_name or not target_amount or not start_date or not end_date:
            messages.error(request, "Please provide all required fields: goal name, target amount, start date, and end date.")
            return redirect('predict_savings')

        try:
            target_amount = Decimal(target_amount)
            start_date = date.fromisoformat(start_date)
            end_date = date.fromisoformat(end_date)
        except Exception as e:
            messages.error(request, "Invalid input format for target amount, start date, or end date.")
            return redirect('predict_savings')

        if end_date <= start_date:
            messages.error(request, "End date must be after the start date.")
            return redirect('predict_savings')

        # Create the goal with end_date
        goal = Goal.objects.create(
            user=user,
            name=goal_name,
            target_amount=target_amount,
            start_date=start_date,
            end_date=end_date  # Add end_date to the Goal model
        )

        # Clean up selected categories
        selected_categories = [category.strip() for category in selected_categories if category.strip()]
        if not selected_categories:
            messages.error(request, "Please select at least one category for exclusion.")
            return redirect('predict_savings')

        # Calculate the number of months between start_date and end_date
        delta = relativedelta(end_date, start_date)
        months_to_goal = delta.years * 12 + delta.months
        if delta.days > 0:  # Add an extra month if there are remaining days
            months_to_goal += 1

        # Calculate expenses excluding selected categories (last 3 months)
        today = date.today()
        three_months_ago = today.replace(month=today.month - 3 if today.month > 3 else 12, day=1)

        expenses = Expense.objects.filter(
            user=user,
            date__gte=three_months_ago
        ).exclude(expensecategory__in=selected_categories)

        total_expenses = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
        avg_monthly_expense = Decimal(total_expenses) / 3 if total_expenses > 0 else Decimal(0)

        # Assume savings potential is 10% of average monthly expenses
        potential_monthly_savings = avg_monthly_expense * Decimal(0.1)

        # Calculate total predicted savings by the end date
        predicted_total_savings = potential_monthly_savings * months_to_goal

        # Calculate suggested monthly savings to reach the target amount
        current_savings = Decimal(goal.current_savings) if goal.current_savings else Decimal(0)
        remaining_to_save = target_amount - current_savings
        suggested_monthly_savings = remaining_to_save / months_to_goal if months_to_goal > 0 else remaining_to_save

        # Render the goal and prediction information
        return render(request, 'predict_savings.html', {
            'goal': goal,
            'total_expenses': round(total_expenses, 2),
            'avg_monthly_expense': round(avg_monthly_expense, 2),
            'months_to_goal': months_to_goal,
            'suggested_monthly_savings': round(suggested_monthly_savings, 2),
            'predicted_total_savings': round(predicted_total_savings, 2),
            'remaining_to_save': round(remaining_to_save, 2),
        })

    # Render the page with the form to set a goal
    return render(request, 'predict_savings.html')


def view_goals(request):
    email = request.session.get('email')  # Get the logged-in user's email from session

    if not email:
        return HttpResponse("User not logged in", status=401)

    try:
        user = user_register.objects.get(email=email)
    except user_register.DoesNotExist:
        return HttpResponse("User not found", status=404)

    # Fetch all goals associated with the logged-in user
    goals = Goal.objects.filter(user=user)
    for goal in goals:
        goal.remaining_amount = goal.target_amount - goal.current_savings

    # Pass the goals to the template
    return render(request, 'view_goals.html', {'goals': goals})
# import re
# from datetime import datetime
# from PyPDF2 import PdfReader
# import spacy
# from django.shortcuts import render, redirect
# from django.contrib import messages
# from .models import Expense, user_register

# # Load spaCy NLP model
# nlp = spacy.load("en_core_web_sm")

# # Define category-related keywords
# CATEGORY_KEYWORDS = {
#     "food": {"restaurant", "cafe", "hotel", "dining", "bar", "bistro", "menu"},
#     "transport": {"uber", "ola", "cab", "bus", "train", "ticket", "metro"},
#     "shopping": {"mall", "store", "shopping", "purchase", "invoice"},
#     "medical": {"hospital", "clinic", "pharmacy", "doctor", "medicine"},
#     "rent": {"rent", "lease", "apartment", "landlord"},
#     "utility": {"electricity", "water bill", "gas", "wifi", "broadband"},
#     "other": set()
# }

# def detect_category(text):
#     """Determine the category based on keywords in the text."""
#     for category, keywords in CATEGORY_KEYWORDS.items():
#         if any(keyword in text for keyword in keywords):
#             return category
#     return "other"

# def extract_amounts(text):
#     """Extract monetary values from the bill text."""
#     amounts = []
#     money_pattern = re.findall(r"₹\s?(\d+[\.,]?\d*)|\b(\d+[\.,]?\d*)\s?(rs|rupees)", text)
#     for match in money_pattern:
#         extracted_value = match[0] or match[1]
#         if extracted_value:
#             amounts.append(float(extracted_value.replace(",", "")))
#     return max(amounts, default=0.0)

# def extract_date(text):
#     """Extract the first detected date from the text."""
#     doc = nlp(text)
#     for ent in doc.ents:
#         if ent.label_ == "DATE":
#             try:
#                 return datetime.strptime(ent.text, "%Y-%m-%d")  # Adjust date format as needed
#             except ValueError:
#                 continue
#     return None

# def process_pdf(request):
#     """Handles the PDF upload and extracts expense details."""
#     email = request.session.get('email')
#     if not email:
#         messages.error(request, "You need to be logged in.")
#         return redirect('login')

#     user = user_register.objects.get(email=email)
#     extracted_data = {"date": None, "category": "other", "amount": None}

#     if request.method == 'POST' and request.FILES.get('pdf'):
#         pdf_file = request.FILES['pdf']
#         reader = PdfReader(pdf_file)
#         text = ""

#         # Extract text from all pages
#         for page in reader.pages:
#             page_text = page.extract_text()
#             if page_text:
#                 text += page_text.lower() + " "

#         # Extract data
#         category = detect_category(text)
#         amount = extract_amounts(text)
#         date = extract_date(text)

#         # Save to database if valid
#         if amount > 0:
#             Expense.objects.create(user=user, expensecategory=category, amount=amount, date=date or datetime.today())
#             extracted_data = {"date": date or datetime.today(), "category": category, "amount": amount}
#             messages.success(request, "Expense saved successfully!")
#         else:
#             messages.error(request, "Failed to extract valid data from the PDF!")

#     return render(request, 'process_pdf.html', {'extracted_data': extracted_data})
from django.db.models import F
from django.contrib import messages

def delete_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id)

    # Find the Income entry for the same user and month
    income_entry = Income.objects.filter(
        user=expense.user,
        date__year=expense.date.year,
        date__month=expense.date.month
    ).first()

    if income_entry:
        # Increase the balance
        income_entry.balance = F('balance') + expense.amount
        income_entry.save(update_fields=['balance'])

    # Delete the expense
    expense.delete()
    messages.success(request, "Expense deleted and balance updated successfully!")

    return redirect('expenselist') 

#pip install -r requirements.txt
# import os
# import easyocr
# import re
# from django.shortcuts import render, redirect
# from django.core.files.storage import default_storage
# from django.conf import settings
# from .models import Expense
# from datetime import datetime

# # Function to extract data from receipt
# def extract_data_from_receipt(image_path):
#     reader = easyocr.Reader(['en'])
#     results = reader.readtext(image_path)

#     extracted_text = " ".join([res[1] for res in results])

#     # Extract amount (e.g., $12.34 or 12.34)
#     amount_match = re.search(r'(\$?(\d{1,5}\.\d{2}))', extracted_text)
#     amount = float(amount_match.group(2)) if amount_match else 0.0

#     # Extract date (e.g., YYYY-MM-DD or MM/DD/YYYY)
#     date_match = re.search(r'(\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4})', extracted_text)
#     date_str = date_match.group(1) if date_match else None

#     # Convert date format if needed (MM/DD/YYYY → YYYY-MM-DD)
#     if date_str and "/" in date_str:
#         date_obj = datetime.strptime(date_str, "%m/%d/%Y")
#         date_str = date_obj.strftime("%Y-%m-%d")

#     # Extract category (based on common words)
#     category_keywords = ['food', 'Transport', 'Groceries', 'Entertainment', 'Shopping', 'Bills']
#     category = next((word for word in category_keywords if word.lower() in extracted_text.lower()), "Other")

#     return {"amount": amount, "date": date_str, "expensecategory": category}

# # View function
# def scan_receipt(request):
#     extracted_data = None
#     email = request.session.get('email')
#     if not email:
#         messages.error(request, "You need to be logged in.")
#         return redirect('login')

#     user = user_register.objects.get(email=email)

#     if request.method == "POST":
#         if 'receipt' in request.FILES:  # If uploading a file
#             receipt = request.FILES['receipt']
#             file_path = default_storage.save(f"receipts/{receipt.name}", receipt)
#             full_path = os.path.join(settings.MEDIA_ROOT, file_path)

#             extracted_data = extract_data_from_receipt(full_path)

#         elif 'save_expense' in request.POST:  # If saving extracted data
#             category = request.POST.get('expensecategory')
#             amount = request.POST.get('amount')
#             date = request.POST.get('date')

#             Expense.objects.create(user=user,expensecategory=category, amount=amount, date=date)
#             return redirect('scan_receipt')

#     return render(request, 'scan_receipt.html', {'extracted_data': extracted_data})

import os
import easyocr
import re
from django.shortcuts import render, redirect
from django.core.files.storage import default_storage
from django.conf import settings
from django.contrib import messages
from .models import Expense
from datetime import datetime

# Function to extract data from receipt
def extract_data_from_receipt(image_path):
    try:
        # Initialize EasyOCR reader for English
        reader = easyocr.Reader(['en'], gpu=False)  # Set gpu=True if you have GPU support
        results = reader.readtext(image_path)

        # Combine all detected text into a single string
        extracted_text = " ".join([res[1] for res in results]).lower()

        # Extract date (e.g., YYYY-MM-DD, MM/DD/YYYY, DD-MM-YYYY)
        date_match = re.search(
            r'(\d{4}-\d{2}-\d{2})|(\d{2}/\d{2}/\d{4})|(\d{2}-\d{2}-\d{4})', 
            extracted_text
        )
        date_str = date_match.group(0) if date_match else None

        # Convert date to YYYY-MM-DD format, default to today if not found
        if date_str:
            if "/" in date_str:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
            elif "-" in date_str:
                date_obj = datetime.strptime(date_str, "%d-%m-%Y") if date_str[2] == "-" else datetime.strptime(date_str, "%Y-%m-%d")
            date_str = date_obj.strftime("%Y-%m-%d")
        else:
            date_str = datetime.now().strftime("%Y-%m-%d")  # Use today's date

        # Extract amount using keywords like "total", "grand total", "amount", "cash"
        amount_keywords = r'(total|grand total|amount|cash|subtotal)[:\s]*\$?\s*(\d{1,5}(?:\.\d{2})?)'
        amount_match = re.search(amount_keywords, extracted_text)
        amount = float(amount_match.group(2)) if amount_match else 0.0

        # Extract category based on keywords
        category_keywords = {
            'Food': ['food', 'restaurant', 'cafe', 'dinner', 'lunch', 'meal'],
            'Transport': ['transport', 'uber', 'taxi', 'bus', 'train', 'fare'],
            'Groceries': ['grocery', 'market', 'store', 'supermarket', 'foodmart'],
            'Entertainment': ['movie', 'cinema', 'event', 'ticket', 'show'],
            'Shopping': ['shop', 'clothing', 'electronics', 'amazon', 'retail'],
            'Bills': ['bill', 'utility', 'electricity', 'water', 'internet']
        }
        category = "Other"
        for cat, keywords in category_keywords.items():
            if any(keyword in extracted_text for keyword in keywords):
                category = cat
                break

        return {
            "amount": amount,
            "date": date_str,
            "expensecategory": category
        }
    except Exception as e:
        print(f"Error extracting data: {e}")
        # Default values if extraction fails
        return {
            "amount": 0.0,
            "date": datetime.now().strftime("%Y-%m-%d"),  # Today's date
            "expensecategory": "Other"
        }

# View function for scanning and saving receipt
def scan_receipt(request):
    extracted_data = None
    email = request.session.get('email')
    
    # Check if user is logged in
    if not email:
        messages.error(request, "You need to be logged in.")
        return redirect('login')

    try:
        user = user_register.objects.get(email=email)
    except user_register.DoesNotExist:
        messages.error(request, "User not found.")
        return redirect('login')

    if request.method == "POST":
        if 'receipt' in request.FILES:  # Handle file upload
            receipt = request.FILES['receipt']
            file_path = default_storage.save(f"receipts/{receipt.name}", receipt)
            full_path = os.path.join(settings.MEDIA_ROOT, file_path)

            # Extract data from the uploaded receipt
            extracted_data = extract_data_from_receipt(full_path)

        elif 'save_expense' in request.POST:  # Handle save action
            category = request.POST.get('expensecategory')
            amount = float(request.POST.get('amount', 0.0))
            date = request.POST.get('date') or datetime.now().strftime("%Y-%m-%d")

            # Save to the database
            if category and amount is not None and date:
                Expense.objects.create(
                    user=user,
                    expensecategory=category,
                    amount=amount,
                    date=date
                )
                messages.success(request, "Expense saved successfully!")
                return redirect('scan_receipt')
            else:
                messages.error(request, "Please fill all fields correctly.")

    return render(request, 'scan_receipt.html', {'extracted_data': extracted_data})
# views.py
import re
from datetime import datetime
from PyPDF2 import PdfReader
import spacy
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Expense, user_register

# Load spaCy NLP model
nlp = spacy.load("en_core_web_sm")

# Enhanced regex patterns
CATEGORY_PATTERN = re.compile(r"(Food|Travel|Shopping|Entertainment|Transport|Bills|Groceries|Other)", re.IGNORECASE)
AMOUNT_PATTERN = re.compile(r"(?i)(Total|Grand\s*Total|Amount\s*Payable|Balance\s*Due|Net\s*Amount|Subtotal|Paid):?\s*(Rs?\.?\s*|\$|€|£)?\s*([\d,]+\.?\d*)")
DATE_PATTERN = re.compile(r"(\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4}|\d{2}-\d{2}-\d{4}|\d{2}\s\w+\s\d{4}|\w+\s\d{1,2},\s\d{4})")

def extract_data_from_text(text):
    """Extracts category, amount, and date from the extracted text with improved accuracy."""
    doc = nlp(text)
    
    # Extract category using both regex and NLP
    category_match = CATEGORY_PATTERN.search(text)
    category = category_match.group(1) if category_match else None
    
    if not category:
        # Fallback to NLP entity recognition
        for ent in doc.ents:
            if ent.label_ in ["PRODUCT", "MONEY", "ORG"]:
                category = ent.text
                break
        category = category or "Other"
    
    # Extract amount with improved logic
    amount_matches = AMOUNT_PATTERN.findall(text)
    amount = 0.0
    if amount_matches:
        # Convert matches to float and take the highest value
        amounts = [float(match[2].replace(",", "")) for match in amount_matches]
        amount = max(amounts) if amounts else 0.0
    
    # Extract date with more format options
    date_match = DATE_PATTERN.search(text)
    date_str = date_match.group(1) if date_match else None
    date = None
    
    if date_str:
        date_formats = [
            "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", 
            "%d %B %Y", "%d-%b-%Y", "%B %d, %Y"
        ]
        for fmt in date_formats:
            try:
                date = datetime.strptime(date_str, fmt).date()
                break
            except ValueError:
                continue
    
    # If no date found, look for date-like entities in NLP
    if not date:
        for ent in doc.ents:
            if ent.label_ == "DATE":
                try:
                    date = datetime.strptime(ent.text, "%B %d, %Y").date()
                    break
                except ValueError:
                    continue
    
    return category.capitalize(), amount, date or datetime.now().date()

def process_pdf(request):
    """Handles PDF upload with improved error handling and feedback."""
    email = request.session.get('email')
    if not email:
        messages.error(request, "Please log in to upload expenses.")
        return redirect('login')

    try:
        user = user_register.objects.get(email=email)
    except user_register.DoesNotExist:
        messages.error(request, "User not found.")
        return redirect('login')

    extracted_data = {"date": None, "category": "Other", "amount": None, "status": "pending"}

    if request.method == 'POST' and request.FILES.get('pdf'):
        try:
            pdf_file = request.FILES['pdf']
            reader = PdfReader(pdf_file)
            text = ""
            
            # Extract text from all pages
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            
            if not text.strip():
                messages.error(request, "No text could be extracted from the PDF.")
                return render(request, 'process_pdf.html', {'extracted_data': extracted_data})

            # Extract and validate data
            category, amount, date = extract_data_from_text(text)
            
            if amount > 0 and category != "Other":
                Expense.objects.create(
                    user=user,
                    expensecategory=category,
                    amount=amount,
                    date=date
                )
                extracted_data = {"date": date, "category": category, "amount": amount, "status": "success"}
                messages.success(request, "Expense saved successfully!")
            else:
                extracted_data = {"date": date, "category": category, "amount": amount, "status": "review"}
                messages.warning(request, "Partial data extracted. Please verify the details.")
                
        except Exception as e:
            messages.error(request, f"Error processing PDF: {str(e)}")
    
    return render(request, 'process_pdf.html', {'extracted_data': extracted_data})